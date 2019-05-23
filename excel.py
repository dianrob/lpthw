from openpyxl import load_workbook
wb = load_workbook('btc.xlsx', data_only = True)

ws = wb.active


values = []

for row in range(2, 460):
    for column in "B":
        cell_name = "{}{}".format(column, row)
        values.append(ws[cell_name].value)

for string in values:
    strings = str(string)
    print strings[slice(5)]
